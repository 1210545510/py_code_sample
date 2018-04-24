# -*- coding: utf-8 -*-

import json
import re
from graphdb_connector import Connector
import config
from config import Fields
from config import NOT_FOUND
from openpyxl import load_workbook
import logic_crawler
import os
import logic_input
import logic_qa

import traceback
import logging
logging.basicConfig()
logger = logging.getLogger("Check")
logger.setLevel(logging.DEBUG)
logger_simple = logging.getLogger("Check_Simple")
logger_simple.setLevel(logging.DEBUG)

def overall(case_name, current_user, full_name, customer, operator):
    logger_simple.info("[Start] Overall Question Logic")

    folder = config.pdf_folder
    final_path = folder + case_name
    config.mkdir(config.answer_folder + case_name)

    result_input = None
    try:
        logger_simple.info("[Start] Format Inputs")
        result_input = logic_input.format_inputs(final_path)
        logger_simple.info("[Finish] Format Inputs")
    except Exception as e:
        tb = traceback.format_exc()
        logger_simple.warning("[Error] Format Inputs")
        logger_simple.warning("[Error] Traceback: " + str(tb))

    logic = Logic(case_name, result_input, current_user, full_name, customer, operator)

    try:
        logic.crawling(operator, case_name)
    except Exception as e:
        tb = traceback.format_exc()
        logger_simple.error("[Error] Crawling")
        logger_simple.error("[Error] Traceback: " + str(tb))

    try:
        logger_simple.info("[Start] A1")
        logic.a1_save()
        logger_simple.info("[Finish] A1")
    except Exception as e:
        tb = traceback.format_exc()
        logger_simple.warning("[Error] A1")
        logger_simple.warning("[Error] Traceback: " + str(tb))

    try:
        logger_simple.info("[Start] A2")
        logic.a2_save()
        logger_simple.info("[Finish] A2")
    except Exception as e:
        tb = traceback.format_exc()
        logger_simple.warning("[Error] A2")
        logger_simple.warning("[Error] Traceback: " + str(tb))


    try:
        logger_simple.info("[Start] B1")
        logic.b1_save()
        logger_simple.info("[Finish] B1")
    except Exception as e:
        tb = traceback.format_exc()
        logger_simple.warning("[Error] B1")
        logger_simple.warning("[Error] Traceback: " + str(tb))

    try:
        logger_simple.info("[Start] B2")
        logic.b2_save()
        logger_simple.info("[Finish] B2")
    except Exception as e:
        tb = traceback.format_exc()
        logger_simple.warning("[Error] B2")
        logger_simple.warning("[Error] Traceback: " + str(tb))


    try:
        logger_simple.info("[Start] B3")
        logic.b3_save()
        logger_simple.info("[Finish] B3")
    except Exception as e:
        tb = traceback.format_exc()
        logger_simple.warning("[Error] B3")
        logger_simple.warning("[Error] Traceback: " + str(tb))


    try:
        logger_simple.info("[Start] B4")
        logic.b4_save()
        logger_simple.info("[Finish] B4")
    except Exception as e:
        tb = traceback.format_exc()
        logger_simple.warning("[Error] B4")
        logger_simple.warning("[Error] Traceback: " + str(tb))

    try:
        logger_simple.info("[Start] B5")
        logic.b5_save()
        logger_simple.info("[Finish] B5")
    except Exception as e:
        tb = traceback.format_exc()
        logger_simple.warning("[Error] B5")
        logger_simple.warning("[Error] Traceback: " + str(tb))

    try:
        logger_simple.info("[Start] B6")
        logic.b6_save()
        logger_simple.info("[Finish] B6")
    except Exception as e:
        tb = traceback.format_exc()
        logger_simple.warning("[Error] B6")
        logger_simple.warning("[Error] Traceback: " + str(tb))

    try:
        logger_simple.info("[Start] B7")
        logic.b7_save()
        logger_simple.info("[Finish] B7")
    except Exception as e:
        tb = traceback.format_exc()
        logger_simple.warning("[Error] B7")
        logger_simple.warning("[Error] Traceback: " + str(tb))

    try:
        logger_simple.info("[Start] B8")
        logic.b8_save()
        logger_simple.info("[Finish] B8")
    except Exception as e:
        tb = traceback.format_exc()
        logger_simple.warning("[Error] B8")
        logger_simple.warning("[Error] Traceback: " + str(tb))

    try:
        logger_simple.info("[Start] B9")
        logic.b9_save()
        logger_simple.info("[Finish] B9")
    except Exception as e:
        tb = traceback.format_exc()
        logger_simple.warning("[Error] B9")
        logger_simple.warning("[Error] Traceback: " + str(tb))

    try:
        logger_simple.info("[Start] B10")
        logic.manual_question('b10')
        logger_simple.info("[Finish] B10")
    except Exception as e:
        tb = traceback.format_exc()
        logger_simple.warning("[Error] B10")
        logger_simple.warning("[Error] Traceback: " + str(tb))

    try:
        logger_simple.info("[Start] B11")
        logic.b11_save()
        logger_simple.info("[Finish] B11")
    except Exception as e:
        tb = traceback.format_exc()
        logger_simple.warning("[Error] B11")
        logger_simple.warning("[Error] Traceback: " + str(tb))

    try:
        logger_simple.info("[Start] B12")
        logic.manual_question('b12')
        logger_simple.info("[Finish] B12")
    except Exception as e:
        tb = traceback.format_exc()
        logger_simple.warning("[Error] B12")
        logger_simple.warning("[Error] Traceback: " + str(tb))

    try:
        logger_simple.info("[Start] B13")
        logic.manual_question('b13')
        logger_simple.info("[Finish] B13")
    except Exception as e:
        tb = traceback.format_exc()
        logger_simple.warning("[Error] B13")
        logger_simple.warning("[Error] Traceback: " + str(tb))

    try:
        logger_simple.info("[Start] B14")
        logic.manual_question('b14')
        logger_simple.info("[Finish] B14")
    except Exception as e:
        tb = traceback.format_exc()
        logger_simple.warning("[Error] B14")
        logger_simple.warning("[Error] Traceback: " + str(tb))

    try:
        logger_simple.info("[Start] B15")
        logic.b15_save()
        logger_simple.info("[Finish] B15")
    except Exception as e:
        tb = traceback.format_exc()
        logger_simple.warning("[Error] B15")
        logger_simple.warning("[Error] Traceback: " + str(tb))

    try:
        logger_simple.info("[Start] B16")
        logic.b16_save()
        logger_simple.info("[Finish] B16")
    except Exception as e:
        tb = traceback.format_exc()
        logger_simple.warning("[Error] B16")
        logger_simple.warning("[Error] Traceback: " + str(tb))

    try:
        logger_simple.info("[Start] B17")
        logic.b17_save()
        logger_simple.info("[Finish] B17")
    except Exception as e:
        tb = traceback.format_exc()
        logger_simple.warning("[Error] B17")
        logger_simple.warning("[Error] Traceback: " + str(tb))

    try:
        logger_simple.info("[Start] B18")
        logic.b18_save()
        logger_simple.info("[Finish] B18")
    except Exception as e:
        tb = traceback.format_exc()
        logger_simple.warning("[Error] B18")
        logger_simple.warning("[Error] Traceback: " + str(tb))

    try:
        logger_simple.info("[Start] B19")
        logic.manual_question('b19')
        logger_simple.info("[Finish] B19")
    except Exception as e:
        tb = traceback.format_exc()
        logger_simple.warning("[Error] B19")
        logger_simple.warning("[Error] Traceback: " + str(tb))

    try:
        logger_simple.info("[Start] B20")
        logic.b20_save()
        logger_simple.info("[Finish] B20")
    except Exception as e:
        tb = traceback.format_exc()
        logger_simple.warning("[Error] B20")
        logger_simple.warning("[Error] Traceback: " + str(tb))

    try:
        logger_simple.info("[Start] B21")
        logic.b21_save()
        logger_simple.info("[Finish] B21")
    except Exception as e:
        tb = traceback.format_exc()
        logger_simple.warning("[Error] B21")
        logger_simple.warning("[Error] Traceback: " + str(tb))

    try:
        logger_simple.info("[Start] B22")
        logic.b22_save()
        logger_simple.info("[Finish] B22")
    except Exception as e:
        tb = traceback.format_exc()
        logger_simple.warning("[Error] B22")
        logger_simple.warning("[Error] Traceback: " + str(tb))

    try:
        logger_simple.info("[Start] B23")
        logic.b23_save()
        logger_simple.info("[Finish] B23")
    except Exception as e:
        tb = traceback.format_exc()
        logger_simple.warning("[Error] B23")
        logger_simple.warning("[Error] Traceback: " + str(tb))

    try:
        logger_simple.info("[Start] B24")
        logic.b24_save()
        logger_simple.info("[Finish] B24")
    except Exception as e:
        tb = traceback.format_exc()
        logger_simple.warning("[Error] B24")
        logger_simple.warning("[Error] Traceback: " + str(tb))

    try:
        logger_simple.info("[Start] B25")
        logic.b25_save()
        logger_simple.info("[Finish] B25")
    except Exception as e:
        tb = traceback.format_exc()
        logger_simple.warning("[Error] B25")
        logger_simple.warning("[Error] Traceback: " + str(tb))

    try:
        logger_simple.info("[Start] B29")
        logic.b29_save()
        logger_simple.info("[Finish] B29")
    except Exception as e:
        tb = traceback.format_exc()
        logger_simple.warning("[Error] B29")
        logger_simple.warning("[Error] Traceback: " + str(tb))

    logger_simple.info("[Finish] Overall Question Logic")



class Logic:

    def __init__(self, case_name, result_input, current_user, full_name, customer, operator):
        self.case_name = case_name
        self.input_fields = result_input
        self.current_user = current_user
        self.full_name = full_name
        self.customer = customer
        self.operator = operator
        self.vessel_crawling = None
        self.company_crawling = None
        self.bloomberg_links = None
        self.customer_dict = None
        self.high_risk_country_dict = None
        self.sanction_country_dict = None
        self.tax_evasion_dict = None
        self.individual_dowjones_crawling = None
        self.individual_bridger_crawling = None
        self.banks_dowjones = self.input_fields["additional banks"]
        self.banks_bridger = self.input_fields["bank of china"]


    def crawling(self, operator, case_name):
        print "Will pinrt case_name======"
        print case_name
        ## Vessel crawling from Lloyds
        vessels = self.input_fields["vessel"]
        # vessels_dict = {item.lower(): item for item in vessels}
        # vessels = vessels_dict.values()
        print "Will print vessel=========="
        print vessels
        dates = self.input_fields["shipment date"]
        self.vessel_crawling, owners = logic_crawler.lloyds_crawling_batch(vessels, dates, operator, case_name)
        print "Owners:", owners
        ## Company crawling from DowJones and BridgerInsight
        self.company_crawling = self.companies_logic(owners, case_name)

        ## Company crawling from Google/Bloomberg
        self.bloomberg_links = logic_qa.logic_bloomberg(self.input_fields, case_name)

        print "print banks_dowjones====="
        print self.banks_dowjones
        print "print banks_bridger====="
        print self.banks_bridger
        ## individula bridgerInsight
        self.individual_bridger_crawling = self.bridger_logic(case_name)

        ## Individual DowJones crawling
        self.individual_dowjones_crawling = self.dowjones_logic(case_name)
        



        # get excel lists
        self.customer_dict = self.load_customer_dict()
        self.high_risk_country_dict = self.load_high_risk_country_dict()
        self.sanction_country_dict = self.load_sanction_country_dict()
        self.tax_evasion_dict = self.load_tax_evasion_dict()
        self.export_dict = self.load_top_10_export_dict()

    def manual_question(self, question_name):
        cntr = Connector()
        logger.info(str(question_name))
        result = {}
        result['answer'] = 'manual'
        result['evidence'] = json.dumps([])
        result["comment"] = ""
        logger.debug("Result of " + question_name + ": " + str(result))
        cntr.saveQuestion(self.case_name, question_name, result)
        logger.info("Save " + question_name + "result: " + str(result))
        return result

    def a1_save(self):
        para = dict()
        para["caseid"] = self.case_name
        customer_dict = self.customer_dict
        result = logic_qa.a1(para, customer_dict)
        cntr = Connector()
        values = dict()
        values["answer"] = result["answer"]
        values["comment"] = result['comment']
        values["evidence"] = json.dumps(result["evidence"])
        if "alert" in result:
            values["alert"] = result["alert"]
        cntr.saveQuestion(self.case_name, "a1", values)
        logger.info("Save A1 result: " + str(values))

    def a2_save(self):
        para = {}
        para["caseid"] = self.case_name
        para[Fields.DOCUMENT_TYPES] = self.input_fields[Fields.DOCUMENT_TYPES]
        customer_dict = self.customer_dict
        high_risk_country_dict = self.high_risk_country_dict
        result = logic_qa.a2(para, customer_dict, high_risk_country_dict)
        output = {}
        cntr = Connector()
        output["answer"] = result["answer"]
        output["comment"] = result["comment"]
        output["evidence"] = json.dumps(result["evidence"])
        if "alert" in result:
            output["alert"] = result["alert"]
        cntr.saveQuestion(self.case_name, "a2", output)
        logger.info("Save A2 result: " + str(output))

    # def b1_save(self):
    #     print "b1_save"
    #     para = {}
    #     para["caseid"] = self.case_name
    #     customer_dict = self.customer_dict
    #     evidences = self.bloomberg_links
    #     cntr = Connector()
    #     result = logic_qa.b1(para, evidences, customer_dict)
    #     output = {}
    #     output["answer"] = result["answer"]
    #     output["comment"] = ""
    #     output["evidence"] = json.dumps(result["evidence"])
    #     cntr.saveQuestion(self.case_name, "b1", output)
    #     logger.info("Save B1 result: " + str(output))
    def b1_save(self):
        # print "b1_save"



        # google_result = self.bloomberg_links
        # result = {}
        # result["answer"] = google_result["answer"]
        # result["comment"] = google_result["comment"]
        # result["evidence"] = json.dumps(google_result["evidence"])

        # # values = {}
        # # # values["answer"] = result["evidence_form"]["answer"]
        # # values["answer"] = ""
        # # values["comment"] = ""
        # # values["evidence"] = []

        # # for description in self.input_fields["invoice info"][0]["commodity info"]:
        # #     values["evidence"].append({"description":description["invoice_description"]})
        # cntr = Connector()

        # # values["evidence"] = json.dumps(values["evidence"])
        # # values["evidence"] = result["evidence_form"]["invoice description"]
        # cntr.saveQuestion(self.case_name, "b1", result)
        # # cntr.saveQuestion(self.case_name, "b1", values)

        # # logger.info("Save B1 result: " + str(values))
        # logger.info("Save B1 result: " + str(result))

        print "b1_save"
        evidences = self.bloomberg_links
        final = {"answer":"manual", "evidence":evidences}
        customer_dict = self.customer_dict


        para = {}

        if len(self.input_fields["invoice info"])!=0:
            para['invoice commodity'] = self.input_fields["invoice info"][0]["commodity info"]
        para['caseid'] = self.case_name
        result = logic_qa.b1(para, customer_dict)

        # json.dumps(company_result["evidence"] + result_v['evidence'])


        if evidences:
            final["answer"] = "manual"
        cntr = Connector()
        values = {"answer": "manual", "comment": "", "evidence": json.dumps(final["evidence"] + result["evidence"])}
        if "alert" in result:
            values["alert"] = result["alert"]
        cntr.saveQuestion(self.case_name, "b1", values)
        logger.info("Save B1 result: " + str(values))

    def b2_save(self):
        para = {}
        para["caseid"] = self.case_name
        para[Fields.DOCUMENT_TYPES] = self.input_fields[Fields.DOCUMENT_TYPES]
        result = logic_qa.b2(para)
        cntr = Connector()
        values = {}
        values["answer"] = result['answer']
        values["comment"] = result["comment"]
        values["evidence"] = json.dumps(result["evidence"])
        if "alert" in result:
            values["alert"] = result["alert"]
        cntr.saveQuestion(self.case_name, "b2", values)
        logger.info("Save B2 result: " + str(values))



    def b3_save(self):
        para = {}
        para["shipment date in application"] = self.input_fields["shipment date in application"]
        para["caseid"] = self.case_name

        para["application date"] = self.input_fields["application date"]
        # print "Will print dates================="
        # print self.input_fields["shipment date in application"]
        # print self.input_fields["application date"]
        result = logic_qa.b3(para)
        cntr = Connector()
        case = cntr.getCase(self.case_name)
        transaction_type = case['transactionType']

        values = {}
        values["answer"] = result["answer"]

        values["comment"] = result["comment"]
        values["evidence"] = []
        values["evidence"] = json.dumps(values["evidence"])
        # print "Will print transaction_type==========="
        # print transaction_type
        if result["answer"] == "na" and (transaction_type == Fields.TYPE_IMPORT_ISSUANCE or transaction_type == Fields.TYPE_IMPORT_DRAWING):
            values["alert"] = "Application date/Shipment date missing."


        cntr.saveQuestion(self.case_name,"b3",values)
        logger.info("Save B2 result: " + str(values))




    def b4_save(self):
        para = {}
        values = {}
        values["answer"] = "na"
        values["comment"] = ""
        values["evidence"] = json.dumps([])
        para['invoice origin'] = self.input_fields["invoice origin"]
        para['certificate origin'] = self.input_fields["certificate origin"]
        # para['invoice commodity'] = self.input_fields["invoice commodity"]
        if len(self.input_fields["invoice info"])!=0:
            para['invoice commodity'] = self.input_fields["invoice info"][0]["commodity info"]
        para["caseid"] = self.case_name
        export_dict = self.export_dict
        result = logic_qa.b4(para, export_dict)
        cntr = Connector()
        values = {}
        values["answer"] = result["answer"]
        values["comment"] = result["comment"]
        values["evidence"] = json.dumps(result["evidence"])
        if "alert" in result:
            values["alert"] = result["alert"]
        cntr.saveQuestion(self.case_name, "b4", values)
        logger.info("Save B4 result: " + str(values))

    @staticmethod
    def currency(currency_unit, amount):
        # TODO: Should be retrieved from other web services
        rate = {
            "usd": 1,
            "USD": 1,
            "CNY": 6.86,
            "EUR": 0.89,
            "JPY": 111.33,
            "GBP": 0.78,
            "KRW": 1118.83,
            "CHF": 0.97
        }
        return amount / rate[currency_unit]

    def b5_save(self):
        if not self.input_fields[Fields.ALIBABA]:
            logger.warning("B5: Product list as Alibaba self.input_fields is empty")
        output_alibaba = []
        alibaba_name_dict = {}
        alibaba_name_price_dict = {}
        for index, item in enumerate(self.input_fields[Fields.ALIBABA]):
            para = dict()
            currency_unit = item["currency"]
            unit_str = item["unit price"]
            min_str = item["min price"]
            max_str = item["max price"]
            para[Fields.INVOICE_UNIT_PRICE] = ""
            para[Fields.INVOICE_MIN_PRICE] = ""
            para[Fields.INVOICE_MAX_PRICE] = ""
            para[Fields.INVOICE_COMMODITY] = item["commodity"]
            para["index"] = index
            if unit_str:
                amount = float(unit_str)
                logger.debug("Start Currency:" + str(currency_unit) + ", price: " + str(amount))
                final_price = Logic.currency(currency_unit, amount)
                logger.debug("Processed Currency: USD, price: " + str(final_price))
                para[Fields.INVOICE_UNIT_PRICE] = str(final_price)
            elif min_str and max_str:
                min_amount = float(min_str)
                max_amount = float(max_str)
                logger.debug("Start Currency:" + str(currency_unit) + ", min: " + str(min_amount) + ", max: " + str(max_amount))
                min_final = Logic.currency(currency_unit, min_amount)
                max_final = Logic.currency(currency_unit, max_amount)
                logger.debug("Processed Currency: USD, min: " + str(min_final) + ", max: " + str(max_final))
                para[Fields.INVOICE_MIN_PRICE] = str(min_final)
                para[Fields.INVOICE_MAX_PRICE] = str(max_final)

            para["unit"] = item["unit"]
            para["current user"] = self.current_user
            para["caseid"] = self.case_name
            
            price = para[Fields.INVOICE_UNIT_PRICE].replace(" ", "")
            if price:
                price = float(price)
                up_limit = float(price * 1.2)
                down_limit = float(price * 0.8)
                name_price_str = item['commodity'] + str(price)
            elif min_amount and max_amount:
                down_limit = float(min_amount)
                up_limit = float(max_amount)
                name_price_str = item['commodity'] + str(up_limit) + str(down_limit)
            else:
                logger.error("No price data found")
                up_limit = down_limit = 0.0
                name_price_str = item['commodity']
            if name_price_str not in alibaba_name_price_dict:
                alibaba_name_price_dict[name_price_str] = True
                invoice_results = {}
                invoice_results['head'] = ['Tolerance min', 'Tolerance max']
                invoice_results['data'] = [[down_limit, up_limit]]
                excel_url = logic_qa.save_excel(para['caseid'], 'b5_' + str(index), invoice_results)
                result = {'answer': 'yes', 'evidence': []}
                result['evidence'].append({'url': excel_url, 'missing': False, 'name': str(para[Fields.INVOICE_COMMODITY]) + ' tolerance max and min', 'type': 'others'})
                output_alibaba.append(result)
            if item["commodity"].lower() in alibaba_name_dict:
                continue
            else:
                alibaba_name_dict[item["commodity"].lower()] = True
                result = logic_qa.b5(para, self.case_name)
                flag = 0
                if len(result["evidence"])==0:
                    continue
                for result_item in output_alibaba:
                    if not result_item["evidence"]:
                        continue
                    if result_item["evidence"][0]["url"] == result["evidence"][0]["url"]:
                        flag = 1
                        logger.info("Same product URL, Don't need to record again!")
                        break
                if flag == 0:
                    output_alibaba.append(result)
        if self.input_fields[Fields.ALIBABA] and len(self.input_fields[Fields.ALIBABA]) > 0:
            result = {"answer": "yes", "evidence": []}
            result['comment'] = '' 
        else:
            result = {"answer": "na", "evidence": []}
            result['comment'] = 'No invoice in the transaction documents' 
	result['alert'] = ""
        for item in output_alibaba:
            print item
            if item["answer"] == "na":
                result["answer"] = "manual"
		result['alert'] = "Unable to find result"
	    elif item['answer'] == 'no' and result['answer'] != 'manual':
                result["answer"] = "manual"
		result['alert'] = "Result out of range"
            if item["evidence"]:
                result["evidence"].extend(item["evidence"])
        print result
        cntr = Connector()
        values = {"answer": result["answer"], "alert": result["alert"], "comment": result['comment'], "evidence": json.dumps(result["evidence"])}
        cntr.saveQuestion(self.case_name, "b5", values)
        logger.info("Save B5 result: " + str(values))


    def b6_save(self):
        para = {}
        para['invoice number'] = self.input_fields["invoice number"]
        para[Fields.DOCUMENT_TYPES] = self.input_fields[Fields.DOCUMENT_TYPES]
        para["caseid"] = self.case_name
        result = logic_qa.b6(para)
        cntr = Connector()
        values = {}
        values["answer"] = result["answer"]
        values["comment"] = result['comment']
        values["evidence"] = json.dumps(result["evidence"])
        if "alert" in result:
            values["alert"] = result["alert"]
        cntr.saveQuestion(self.case_name, "b6", values)
        logger.info("Save B6 result: " + str(values))


    ##new
    def b7_save(self):
        para = {}
        para["lading info"] = self.input_fields["lading master"]
        para["invoice info"] = self.input_fields["invoice info"]
        para["pl gross info"] = self.input_fields["pl gross weight"]
        para["pl net weight info"] = self.input_fields["pl net weight"]
        para["document type"] = self.input_fields["document types"]
        para["caseid"] = self.case_name

        values = {}
        values["answer"] = "manual"
        values["comment"] = ""  
        values["evidence"] = []

        result = logic_qa.b7(para)
        values["evidence"] = result["evidence"]
        cntr = Connector()
        # # values = {"answer": result["answer"], "comment": "", "evidence": json.dumps(result["evidence"])}

        # values["answer"] = result["answer"]

        values["evidence"] = json.dumps(values["evidence"])
        if "alert" in result:
            values["alert"] = result["alert"]
        cntr.saveQuestion(self.case_name, "b7", values)

        logger.info("Save B7 result: " + str(values))




    def b8_save(self):
        para={}
        para['invoice unit price'] = []
        para['invoice commodity'] = []
        para[Fields.DOCUMENT_TYPES] = self.input_fields[Fields.DOCUMENT_TYPES]
        para['invoice number'] = self.input_fields["invoice number"]
        for item in self.input_fields[Fields.ALIBABA]:
            para['invoice unit price'].append(item["unit price"])
            para['invoice commodity'].append(item["commodity"])
        result = logic_qa.b8(para)
        cntr = Connector()
        values = {}
        values["answer"] = result["answer"]
        values["comment"] = result['comment']
        values["evidence"] = json.dumps(result["evidence"])
        if "alert" in result:
            values["alert"] = result["alert"]
        cntr.saveQuestion(self.case_name, "b8", values)
        logger.info("Save B8 result: " + str(values))

    def b9_save(self):
        para = {}
        para["caseid"] = self.case_name
        para[Fields.DOCUMENT_TYPES] = self.input_fields[Fields.DOCUMENT_TYPES]
        result = logic_qa.b9(para)
        output = {}
        cntr = Connector()
        output["answer"] = result["answer"]
        output["comment"] = result["comment"]
        output["evidence"] = json.dumps(result["evidence"])
        if "alert" in result:
            output["alert"] = result["alert"]
        cntr.saveQuestion(self.case_name, "b9", output)
        logger.info("Save B9 result: " + str(output))


    def b11_save(self):
        # if len(self.input_fields['invoice description']) == 0 or len(self.input_fields['lading description']) == 0 or len(self.input_fields['invoice gross weight'][0])==0 or\
        #         len(self.input_fields['invoice quantity'])==0 or len(self.input_fields['lading gross weight']) == 0 or len(self.input_fields['lading quantity']) == 0 or\
        #         len(self.input_fields['lading net weight'])==0  or len(self.input_fields['invoice net weight'])==0 or len(self.input_fields['lading net weight']) == 0 or\
        #         len(self.input_fields["invoice unit price"])==0 or len(self.input_fields["lading unit price"])==0 or len(self.input_fields["invoice unit price"]):
        #     result = dict()
        #     result['evidence'] = []
        #     result['answer'] = "na"
        #     fileObject = open(config.answer_folder+self.case_name+"/b11.json", 'w')
        #     fileObject.write(json.dumps(result))
        #     fileObject.close()
        #     return
        # para={}
        # para['invoice gross weight'] = self.input_fields['invoice gross weight'][0]
        # para['invoice quantity'] = self.input_fields['invoice quantity'][0]
        # para['lading gross weight'] = self.input_fields['lading gross weight'][0]
        # para['lading quantity'] = self.input_fields['lading quantity'][0]
        # para['invoice description'] = self.input_fields['invoice description'][0]
        # para['lading description'] = self.input_fields['lading description'][0]
        # para['invoice net weight'] = self.input_fields['invoice net weight'][0]
        # para['lading net weight'] = self.input_fields['lading net weight'][0]
        # para["invoice unit price"] = self.input_fields["invoice unit price"][0]
        # para["lading unit price"] = self.input_fields["lading unit price"][0]
        # para["invoice value"] = self.input_fields["invoice unit price"][0]
        # para["lading value"] = self.input_fields["invoice unit price"][0]
        # result = logic_qa.b11(para)
        # cntr = Connector()
        # values = {}
        # values["answer"] = result["answer"]
        # values["comment"] = ""
        # values["evidence"] = json.dumps(result["evidence"])
        # cntr.saveQuestion(self.case_name, "b11", values)
        # logger.info("Save B11 result: " + str(values))
        para = {}
        para["lading info"] = self.input_fields["lading master"]
        para["invoice info"] = self.input_fields["invoice info"]
        para["pl gross info"] = self.input_fields["pl gross weight"]
        para["pl net weight info"] = self.input_fields["pl net weight"]
        para["document type"] = self.input_fields["document types"]
        para["caseid"] = self.case_name

        values = {}
        values["answer"] = "manual"
        values["comment"] = ""
        values["evidence"] = []

        result = logic_qa.b11(para)
        values["evidence"] = result["evidence"]
        cntr = Connector()
        # # values = {"answer": result["answer"], "comment": "", "evidence": json.dumps(result["evidence"])}

        # values["answer"] = result["answer"]

        values["evidence"] = json.dumps(values["evidence"])
        if "alert" in result:
            values["alert"] = result["alert"]
        cntr.saveQuestion(self.case_name, "b11", values)

        logger.info("Save B11 result: " + str(values))

    def b15_save(self):
        para = {}
        para[Fields.INVOICE_ORIGIN] = self.input_fields[Fields.INVOICE_ORIGIN]
        para[Fields.CERTIFICATE_ORIGIN] = self.input_fields[Fields.CERTIFICATE_ORIGIN]
        para[Fields.LC_PL_COUNTRY] = self.input_fields[Fields.LC_PL_COUNTRY]
        para[Fields.BL_PL_COUNTRY] = self.input_fields[Fields.BL_PL_COUNTRY]
        para[Fields.TRUCK_PL_COUNTRY] = self.input_fields[Fields.TRUCK_PL_COUNTRY]
        result = logic_qa.b15(para)
        cntr = Connector()
        values = {}
        values["answer"] = result["answer"]
        values["comment"] = ""
        values["evidence"] = json.dumps(result["evidence"])
        cntr.saveQuestion(self.case_name, "b15", values)
        logger.info("Save B15 result: " + str(values))

    def b16_save(self):
        para = {}
        para["caseid"] = self.case_name
        para[Fields.INVOICE_ORIGIN] = [item for item in self.input_fields[Fields.INVOICE_ORIGIN] if item]
        para[Fields.CERTIFICATE_ORIGIN] = [item for item in self.input_fields[Fields.CERTIFICATE_ORIGIN] if item]
        para[Fields.LC_PL_COUNTRY] = [item for item in self.input_fields[Fields.LC_PL_COUNTRY] if item]
        para[Fields.LC_PD_COUNTRY] = [item for item in self.input_fields[Fields.LC_PD_COUNTRY] if item]
        para[Fields.BL_PL_COUNTRY] = [item for item in self.input_fields[Fields.BL_PL_COUNTRY] if item]
        para[Fields.BL_PD_COUNTRY] = [item for item in self.input_fields[Fields.BL_PD_COUNTRY] if item]
        para[Fields.TRUCK_PL_COUNTRY] = [item for item in self.input_fields[Fields.TRUCK_PL_COUNTRY] if item]
        para[Fields.TRUCK_PD_COUNTRY] = [item for item in self.input_fields[Fields.TRUCK_PD_COUNTRY] if item]
        para[Fields.AIRWAY_DEPARTURE_COUNTRY] = [item for item in self.input_fields[Fields.AIRWAY_DEPARTURE_COUNTRY] if item]
        para[Fields.AIRWAY_ARRIVAL_COUNTRY] = [item for item in self.input_fields[Fields.AIRWAY_ARRIVAL_COUNTRY] if item]
        para[Fields.BL_TRANSSHIPMENT_COUNTRY] = [item for item in self.input_fields[Fields.BL_TRANSSHIPMENT_COUNTRY] if item]
        para[Fields.DOCUMENT_TYPES] = self.input_fields[Fields.DOCUMENT_TYPES]
        sanction_country_dict = self.sanction_country_dict
        result_c = self.company_crawling
        result_v = self.vessels_logic("b16")
        result_d = self.individual_dowjones_crawling
        result_b = self.individual_bridger_crawling
        result = logic_qa.b16(para, sanction_country_dict)
        result["comment"] = json.dumps(result["comment"])
        result["evidence"] = result_c["evidence"] + result_d["evidence"] + result_v["evidence"] + result_b["evidence"] + result['evidence']
        sanction_country_url = self.get_sanction_country_url()
        result['evidence'].append({'url': sanction_country_url, 'missing': False, 'name': 'Sanction Countries', 'type': 'others'})
        result['evidence'] = json.dumps(result['evidence'])
        cntr = Connector()
        cntr.saveQuestion(self.case_name, "b16", result)
        logger.info("Save B16 result: " + str(result))

    def b17_save(self):
        para = {}
        para["caseid"] = self.case_name
        result = logic_qa.b17(para)
        output = {}
        cntr = Connector()
        output["answer"] = result["answer"]
        output["comment"] = result["comment"]
        output["evidence"] = json.dumps(result["evidence"])
        cntr.saveQuestion(self.case_name, "b17", output)
        logger.info("Save B17 result: " + str(output))
        return result

    def b18_save(self):
        para = {}
        para["caseid"] = self.case_name
        para[Fields.DOCUMENT_TYPES] = self.input_fields[Fields.DOCUMENT_TYPES]
        result = logic_qa.b18(para)
        output = {}
        cntr = Connector()
        output["answer"] = result["answer"]
        output["comment"] = result["comment"]
        output["evidence"] = json.dumps(result["evidence"])
        cntr.saveQuestion(self.case_name, "b18", output)
        logger.info("Save B18 result: " + str(output))

    def b20_save(self):
        para = {}
        para["caseid"] = self.case_name
        high_risk_country_dict = self.high_risk_country_dict
        para[Fields.INVOICE_ORIGIN] = [item for item in self.input_fields[Fields.INVOICE_ORIGIN] if item]
        para[Fields.CERTIFICATE_ORIGIN] = [item for item in self.input_fields[Fields.CERTIFICATE_ORIGIN] if item]
        para[Fields.LC_PL_COUNTRY] = [item for item in self.input_fields[Fields.LC_PL_COUNTRY] if item]
        para[Fields.LC_PD_COUNTRY] = [item for item in self.input_fields[Fields.LC_PD_COUNTRY] if item]
        para[Fields.BL_PL_COUNTRY] = [item for item in self.input_fields[Fields.BL_PL_COUNTRY] if item]
        para[Fields.BL_PD_COUNTRY] = [item for item in self.input_fields[Fields.BL_PD_COUNTRY] if item]
        para[Fields.TRUCK_PL_COUNTRY] = [item for item in self.input_fields[Fields.TRUCK_PL_COUNTRY] if item]
        para[Fields.TRUCK_PD_COUNTRY] = [item for item in self.input_fields[Fields.TRUCK_PD_COUNTRY] if item]
        para[Fields.AIRWAY_DEPARTURE_COUNTRY] = [item for item in self.input_fields[Fields.AIRWAY_DEPARTURE_COUNTRY] if item]
        para[Fields.AIRWAY_ARRIVAL_COUNTRY] = [item for item in self.input_fields[Fields.AIRWAY_ARRIVAL_COUNTRY] if item]
        para[Fields.BL_TRANSSHIPMENT_COUNTRY] = [item for item in self.input_fields[Fields.BL_TRANSSHIPMENT_COUNTRY] if item]
        para[Fields.DOCUMENT_TYPES] = self.input_fields[Fields.DOCUMENT_TYPES]
        result = logic_qa.b20(para, high_risk_country_dict)
        result_v = self.vessels_logic("b20")
        result["evidence"] = json.dumps(result_v["evidence"] + result["evidence"])
        cntr = Connector()
        cntr.saveQuestion(self.case_name, "b20", result)
        logger.info("Save B20 result: " + str(result))

    def b21_save(self):
        para = {}
        para['invoice buyer'] = self.input_fields["invoice buyer"]
        para['invoice seller'] = self.input_fields["invoice seller"]
        para["caseid"] = self.case_name
        tax_evasion_dict = self.tax_evasion_dict
        result = logic_qa.b21(para, tax_evasion_dict)
        tax_heaven_url = self.get_tax_heaven_url()
        result['evidence'].append({'url': tax_heaven_url, 'missing': False, 'name': 'Tax Heavens', 'type': 'others'})
        result["evidence"] = json.dumps(result["evidence"])
        result["comment"] = ""
        cntr = Connector()
        cntr.saveQuestion(self.case_name, "b21", result)
        logger.info("Save B21 result: " + str(result))

    def b22_save(self):
        result_v = self.vessels_logic("b22")
        result = dict()
        result["answer"] = "manual"
        result["comment"] = ""
        result["evidence"] = json.dumps(result_v["evidence"])
        cntr = Connector()
        cntr.saveQuestion(self.case_name, "b22", result)
        logger.info("Save B22 result: " + str(result))


    def b23_save(self):
        result_v = self.vessels_logic("b23")
        result = dict()
        result["answer"] = "manual"
        result["comment"] = ""
        result["evidence"] = json.dumps(result_v["evidence"])
        cntr = Connector()
        cntr.saveQuestion(self.case_name, "b23", result)
        logger.info("Save B23 result: " + str(result))


    def b24_save(self):
        result_v = self.vessels_logic("b24")
        result_d = self.individual_dowjones_crawling
        result_b = self.individual_bridger_crawling
        company_result = self.company_crawling # self.companies_logic()
        # reverse the answer according to the question
#        if company_result["answer"] == 'yes':
#            company_result["answer"] = 'no'
#        else:
#            company_result["answer"] = 'yes'
        result = {"answer": 'manual', "comment": company_result["comment"]}
        result["evidence"] = company_result["evidence"] + result_v['evidence'] + result_b["evidence"] + result_d["evidence"]
        flag = False
        for evidence in result["evidence"]:
            if 'detail' in evidence:

                for detail_item in evidence['detail']:
                    if 'score' in detail_item and int(detail_item['score']) == 100:
                        result['answer'] = 'no'
                        # result['alert'] = '100% hit'
                        flag = True
                        break
                if flag:
                    break
            if 'isHit' in evidence and evidence['isHit']:
                result['answer'] = 'no'
                # result['alert'] = '100% hit'
                flag = True
            if flag:
                break
        result['answer'] = 'manual'
        result['evidence'] = json.dumps(result['evidence'])
        logger.info("B24 evidence" + str(result["evidence"]))
        cntr = Connector()
        cntr.saveQuestion(self.case_name, "b24", result)
        logger.info("Save B24 result: " + str(result))


    def b25_save(self):
        company_result = self.company_crawling  # self.companies_logic()
        result_d = self.individual_dowjones_crawling
        result_b = self.individual_bridger_crawling
        # result = {"answer": company_result["answer"], "comment": company_result["comment"]}
        result = {"answer": 'manual', "comment": company_result["comment"]}
        pre_evidence_list = company_result["evidence"] + result_b["evidence"]
        new_evidence_list = []
        for evidence in pre_evidence_list:
            if evidence['type'] == 'bridgerInsight':
                new_evidence_list.append(evidence)
        result["evidence"] = json.dumps(new_evidence_list)
        cntr = Connector()
        cntr.saveQuestion(self.case_name, "b25", result)
        logger.info("Save B25 result: " + str(result))

    def b29_save(self):
        para = {}
        para["caseid"] = self.case_name
        result = logic_qa.b29(para)
        output = {}
        cntr = Connector()
        output["answer"] = result["answer"]
        output["comment"] = result["comment"]
        output["evidence"] = json.dumps(result["evidence"])
        if "alert" in result:
            output["alert"] = result["alert"]
        cntr.saveQuestion(self.case_name, "b29", output)
        logger.info("Save B29 result: " + str(output))

    def vessels_logic(self, question):
        output = list()
        question = question.lower()
        total_vessels = self.input_fields["vessel"]
        total_date = self.input_fields["shipment date"]
        print "Will print total_vessels======"
        print total_vessels
        if question == "b16":
            for i in range(0, len(total_vessels)):
                para = dict()
                para["invoice origin"] = self.input_fields["invoice origin"]
                para["lading country"] = self.input_fields["lading country"]
                para["lading vessel"] = total_vessels[i]
                para["date"] = total_date[i]
                para["date"] = para["date"].replace(" ", "")
                result = logic_qa.logic_vessel(para, question, self.vessel_crawling[i]) # logic_qa.b16(para)
                output.append(result)
        elif question == "b20" or question == "b24" or question == "b16" or question == "b22":
            for i in range(0, len(total_vessels)):
                para = dict()
                para["lading port of loading"] = self.input_fields["lading port of loading"]
                para["lading port of discharge"] = self.input_fields["lading port of discharge"]
                para['invoice origin'] = self.input_fields["invoice origin"]
                para["certificate origin"] = self.input_fields["certificate origin"]
                para["lading vessel"] = total_vessels[i]
                para["date"] = total_date[i]
                para["date"] = para["date"].replace(" ", "")
                result = logic_qa.logic_vessel(para, question, self.vessel_crawling[i]) # logic_qa.b20(para)
                output.append(result)
        elif question == "b21":
            for i in range(0, len(total_vessels)):
                para = dict()
                para["invoice origin"] = self.input_fields["invoice origin"]
                para["lading country"] = self.input_fields["lading country"]
                para["lading vessel"] = total_vessels[i]
                para["date"] = total_date[i]
                para["date"] = para["date"].replace(" ", "")
                result = logic_qa.logic_vessel(para, question, self.vessel_crawling[i]) # logic_qa.b21(para)
                output.append(result)
#        elif question == "b22":
#            for i in range(0, len(total_vessels)):
#                para = dict()
#                para["lading vessel"] = total_vessels[i]
#                para["date"] = total_date[i]
#                para["date"] = para["date"].replace(" ", "")
#                result = logic_qa.logic_vessel(para, question, self.vessel_crawling[i]) # logic_qa.b22(para)
#                output.append(result)
        elif question == "b23":
            for i in range(0, len(total_vessels)):
                para = dict()
                para["lading vessel"] = total_vessels[i]
                para["date"] = total_date[i]
                para["date"] = para["date"].replace(" ", "")
                if len(self.input_fields["invoice description"]) == 0:
                    para['invoice description'] = ""
                else:
                    para['invoice description'] = self.input_fields["invoice description"][0]
                if len(self.input_fields["lading gross weight"]) == 0:
                    para["lading gross weight"] = 0
                else:
                    para["lading gross weight"] = self.input_fields["lading gross weight"][0]
                if len(self.input_fields["lading description"]) == 0:
                    para['lading description'] = ""
                else:
                    para['lading description'] = self.input_fields["lading description"][0]
                result = logic_qa.logic_vessel(para, question, self.vessel_crawling[i]) # logic_qa.b23(para)
                output.append(result)

        values = {"answer": "no", "comment": "", "evidence": []}

        ### Set the answer
        data = output
        print "print date from lloyds=========="
        print data
        result_list = {"answer": "no"}
        evidence = []
        for item in data:
            if 'missing' in item and item['missing'] == True:
                # evidence += item['evidence'] + '; '
                values["evidence"].append({'url': config.NOT_FOUND, 'type': 'lloyds', 'name': item['evidence'], 'detail': [], 'missing': True})

            if item["answer"] == "yes":
                values["answer"] = "yes"
            if item["evidence"]:
                values["evidence"].append(item["evidence"][0])

        ### Add DowJones and BridgerInsight Result for Vessel
        # for vessel_name in total_vessels:

        #     dowjones_res = logic_crawler.dowjones_crawling(vessel_name, self.operator, self.case_name)
            
        #     if dowjones_res != NOT_FOUND:
        #         detail_lists = []
        #         lis = []
        #         for res in dowjones_res:
        #             detail_lists.append(res)

        #         # detail_lists = sorted(detail_lists, key=str.lower)
        #         for li in detail_lists:
        #             company_name = li.split("_")
        #             post_word = company_name[len(company_name)-1]
        #             item = None
        #             if post_word != 'summary.pdf':
        #                 item = {'company_name': company_name[len(company_name)-1], 'list_url': li}
        #                 lis.append(item)
        #         # for res in dowjones_res:
        #         values["evidence"].append({'url': detail_lists[0], 'type': 'dowjones', 'name': vessel_name, 'detail': lis})

        #         # detail_list = []
        #         # for res in dowjones_res:
        #         #     detail_list.append(res)

        #         # for res in dowjones_res:
        #         #     values["evidence"].append({'url': str(res), 'type': 'dowjones', 'name': vessel_name, 'detail': detail_list})

        #     config.mkdirs(config.local_bridge_folder + config.date_format(config.get_today()))
        #     path = logic_crawler.bridger_pdf_local(vessel_name)
        #     if not os.path.exists(path):
        #         result = logic_crawler.bridger_crawling(vessel_name, self.full_name, self.operator, self.case_name)
        #         if result != NOT_FOUND:
        #             values["evidence"].append(result)
        #     else:
        #         return_url = config.current_url + logic_crawler.bridger_pdf_web(vessel_name)
        #         ev = {"type": "bridgerInsight", "url": return_url, 'name': vessel_name}
        #         values["evidence"].append(ev)
        print "Will print lloyds values========"
        print values
        return values

    def get_sanction_country_url(self):
        return config.current_url + config.web_excel_folder + 'sanction_country.xlsx'

    def get_tax_heaven_url(self):
        return config.current_url + config.web_excel_folder + 'common_tax_heavens.xlsx'

    def update_customer_dict(self, name, customer_dict):
        wb_url = 'app/data/excel/customer_' + name + '.xlsx'
        wb = load_workbook(filename=wb_url)
        ws = wb.active
        customer_attr_list = ['name', 'id', 'area', 'volume']
        for i, row in enumerate(ws.iter_rows()):
            if i == 0 or i == 1:
                continue
            else:
                customer_item = {}
                for j, cell in enumerate(row):
                    if cell.value == None:
                        customer_item[customer_attr_list[j]] = 'n/a'
                    else:
                        customer_item[customer_attr_list[j]] = unicode(cell.value)
                customer_dict[customer_item['id']] = customer_item
        return customer_dict

    def load_customer_dict(self):
        customer_dict = {}
        self.update_customer_dict('ny', customer_dict)
        self.update_customer_dict('chicago', customer_dict)
        self.update_customer_dict('la', customer_dict)
        return customer_dict

    def load_high_risk_country_dict(self):
        wb_url = './app/data/excel/country_abbr_risk.xlsx'
        wb = load_workbook(filename=wb_url)
        ws = wb.active
        high_risk_country_dict = {}
        for i, row in enumerate(ws.iter_rows()):
            if i == 0:
                continue
            else:
                if row[2].value == 'high':
                    if row[0].value:
                        high_risk_country_dict[row[0].value.upper()] = True
                    if row[1].value:
                        high_risk_country_dict[row[1].value.upper()] = True
        return high_risk_country_dict

    def load_tax_evasion_dict(self):
        tax_evasion_dict = {}
        wb_url = './app/data/excel/country_abbr_risk.xlsx'
        wb = load_workbook(filename=wb_url)
        ws = wb.active
        return tax_evasion_dict

    def load_top_10_export_dict(self):
        export_dict = {}
        wb_url = './app/data/excel/top_10_exports.xlsx'
        wb = load_workbook(filename=wb_url)
        ws = wb.active
        for row in ws.iter_rows():
            if row[0].value:
                export_dict[row[0].value.lower()] = row[2].value
            if row[1].value:
                export_dict[row[1].value.lower()] = row[2].value
        return export_dict


        return tax_evasion_dict


    def load_sanction_country_dict(self):
        wb_url = './app/data/excel/sanction_country.xlsx'
        wb = load_workbook(filename=wb_url)
        ws = wb.active
        sanction_country_dict = {}
        for i, row in enumerate(ws.iter_rows()):
            if i == 0:
                continue
            else:
                if row[0].value:
                    sanction_country_dict[row[0].value] = True
                if row[1].value:
                    sanction_country_dict[row[1].value] = True
        return sanction_country_dict
    def bridger_logic(self, case_name):
        banks = {}
        print "Will print all bank of china============"
        print self.banks_bridger
        # banks.update({item.lower():item for item in self.input_fields["lc advising bank"]})
        # banks.update({item.lower():item for item in self.input_fields["cl correspondent bank"]})
        # banks.update({item.lower():item for item in self.input_fields["application transfer bank"]})
        # banks.update({item.lower():item for item in self.input_fields["application assignee bank"]})
        if len(self.banks_bridger) != 0:
            banks.update({item:item for item in self.banks_bridger})
        result = logic_qa.logic_bridger(banks, self.full_name, self.operator, case_name)
        values = dict()
        values["answer"] = result["answer"]
        values["comment"] = ""
        values["evidence"] = result["evidence"]
        print "result evidence ==="
        print result["evidence"]
        return values

    def dowjones_logic(self, case_name):
        banks = {}
        print "Will print all additoinal bank================"
        print self.banks_dowjones
        if len(self.banks_dowjones) != 0:
            banks.update({item:item for item in self.banks_dowjones})
        result = logic_qa.logic_dowjones(banks, self.full_name, self.operator, case_name)
        values = dict()
        values["answer"] = result["answer"]
        values["comment"] = ""
        values["evidence"] = result["evidence"]
        print "result evidence ==="
        print result["evidence"]
        return values



    def companies_logic(self, owners, case_name):
        companies = {}
        if owners != NOT_FOUND:
            companies.update({item.lower():item for item in owners})
        companies.update({item.lower():item for item in self.input_fields["invoice buyer"]})
        companies.update({item.lower():item for item in self.input_fields["invoice seller"]})
        companies.update({item.lower():item for item in self.input_fields["lading shipper"]})
        companies.update({item.lower():item for item in self.input_fields["lading notify party"]})
        companies.update({item.lower():item for item in self.input_fields["lading consignee"]})
        companies.update({item.lower():item for item in self.input_fields["third party"]})
        companies.update({item.lower():item for item in self.input_fields["vessel"]})
        companies.update({item.lower():item for item in self.input_fields["lc applicant"]})
        companies.update({item.lower():item for item in self.input_fields["lc beneficiary"]})
        companies.update({item.lower():item for item in self.input_fields["lc end user"]})
        companies.update({item.lower():item for item in self.input_fields["master"]})

        ##Added 10/16
        companies.update({item.lower():item for item in self.input_fields["invoice third party"]})
        companies.update({item.lower():item for item in self.input_fields["bl third party"]})
        companies.update({item.lower():item for item in self.input_fields["pl third party"]})
        companies.update({item.lower():item for item in self.input_fields["application third party"]})
        companies.update({item.lower():item for item in self.input_fields["cl third party"]})
        companies.update({item.lower():item for item in self.input_fields["certificate third party"]})
        companies.update({item.lower():item for item in self.input_fields["lc third party"]})
        companies.update({item.lower():item for item in self.input_fields["insurance third party"]})
        companies.update({item.lower():item for item in self.input_fields["airway third party"]})
        companies.update({item.lower():item for item in self.input_fields["truck third party"]})
        companies.update({item.lower():item for item in self.input_fields["draft third party"]})
        companies.update({item.lower():item for item in self.input_fields["application buyer"]})
        companies.update({item.lower():item for item in self.input_fields["application seller"]})
        companies.update({item.lower():item for item in self.input_fields["application second beneficiary"]})
        companies.update({item.lower():item for item in self.input_fields["application assignee"]})
        companies.update({item.lower():item for item in self.input_fields["insurance company"]})
        companies.update({item.lower():item for item in self.input_fields["airway shipper"]})
        companies.update({item.lower():item for item in self.input_fields["airway receiver"]})
        companies.update({item.lower():item for item in self.input_fields["truck company name"]})
        companies.update({item.lower():item for item in self.input_fields["truck shipper"]})
        companies.update({item.lower():item for item in self.input_fields["truck receiver"]})
        companies = companies.values()

        result = logic_qa.logic_companies(companies, self.full_name, self.operator, case_name)
        values = dict()
        values["answer"] = result["answer"]
        values["comment"] = ""
        values["evidence"] = result["evidence"]
        print "result evidence ==="
        print result["evidence"]
        return values

def load_json(case_name, question_name):
    with open(config.answer_folder + case_name + "/" + question_name + ".json") as data_file:
        data = json.load(data_file)
        print str(data)
    return data

if __name__ == "__main__":
    case_name = "20170321170117"
    current_user = "benjie"
    full_name = "huluwa"
    customer = "benjie"

    # overall(case_name, current_user, full_name, customer, operator)
    logic = Logic(case_name, None, current_user, full_name, customer, operator)
    print logic.load_customer_dict()
    print logic.load_high_risk_country_dict()
    print logic.load_sanction_country_dict()
